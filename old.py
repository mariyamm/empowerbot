from flask import Flask, request, jsonify, render_template
from google.cloud import aiplatform
from sentence_transformers import SentenceTransformer

import openpyxl
from pinecone import Pinecone, ServerlessSpec # Example vector database, replace if needed
from langchain_text_splitters import MarkdownHeaderTextSplitter
from transformers import pipeline  # For summarization
import faiss
import numpy as np
from google.oauth2 import service_account

from langchain_google_vertexai import VertexAI
from langchain_core.runnables import RunnableLambda
from langchain_core.messages import AIMessage
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain, RetrievalQA
from langchain.memory import ConversationBufferMemory
from langchain_pinecone import PineconeVectorStore
from langchain_pinecone import PineconeEmbeddings

import openpyxl
import os
import re

import time
import requests
from requests.exceptions import HTTPError
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from google.api_core.exceptions import ResourceExhausted

from datetime import datetime


from managers.google_cloud_auth import GoogleCloudAuth
from managers.pinecone_manager import PineconeManager
from managers.knowledge_base_manager import KnowledgeBaseManager
from managers.chatbot import Chatbot
from managers.user_profile_manager import UserProfileManager

MODEL_ID = "gemini-pro"  # Replace with your actual endpoint ID
PROJECT_ID = "grantgpt-433615"
REGION = "us-central1"



# 0. AUTHENTICATE WITH GOOGLE CLOUD PLATFORM
def authicate_google():
    
    key_path = "C:/Users/a884470/backend/grantgp.json"
    
    credentials = service_account.Credentials.from_service_account_file(key_path)


    aiplatform.init(
        project=PROJECT_ID,
        credentials=credentials,
        location=REGION
    )
    print("Authenticated with Google Cloud Platform")


# 1. SET UP VERTEX AI ENVIRONMENT
authicate_google()

# 2. LOAD PINECONE MODEL FOR EMBEDDING
embeddings = SentenceTransformer('all-MiniLM-L6-v2')

# 2A DEFINE THE EMBEDDING DIMENSION (EXAMPLE: 384 FOR 'ALL-MINILM-L6-V2')
embedding_dimension = embeddings.get_sentence_embedding_dimension()


# 3. INITIALIZE PINECONE VECTOR DATABASE
pc = Pinecone(api_key="pcsk_7RJrPF_3YGAhAWVnZ4zNTiU65AjkVE83kjJHborJTmLCGq13xt1xmred7smfcqrzGJQ1ME")
cloud = os.environ.get('PINECONE_CLOUD') or 'aws'
region = os.environ.get('PINECONE_REGION') or 'us-east-1'
spec = ServerlessSpec(cloud=cloud, region=region)

# 4. INITIALIZE TWO SEPARATE INDEXES FOR FULL DOCUMENTS AND SUMMARIES
# Helper function to create an index
def create_index_if_not_exists(index_name):
    if index_name not in pc.list_indexes().names():
        pc.create_index(
            name=index_name,
            dimension=embedding_dimension,
            metric="cosine",
            spec=spec
        )
        while not pc.describe_index(index_name).status['ready']:
            time.sleep(1)
    return pc.Index(index_name)
index_name_full = "life-coaches-full"
index_name_summary = "life-coaches-summary"
index_full = create_index_if_not_exists(index_name_full)
index_summary = create_index_if_not_exists(index_name_summary)


# 5. INITIALIZE SUMMARIZATION MODEL
summarizer = pipeline("summarization")


# 6. LOAD PREVIOUSLY EMBEDDED KNOWLEDGE BASE
def load_knowledge_base():
  
   # Load the Excel workbook
    file_path = r"C:\Users\a884470\empowerbot-env\LifeCoaches.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    
    
    # Iterate over coaches (starting from row 2)
    for row in range(2, sheet.max_row + 1):
        coach_name = sheet[f"A{row}"].value  # Get coach name from column A

        # Iterate over topics (starting from column 4, since A, B, and C are reserved)
        for col in range(4, sheet.max_column + 1):
            topic = sheet.cell(row=1, column=col).value  # Get topic from row 1
            if topic:  # Ensure there is a topic
                existing_teaching = sheet.cell(row=row, column=col).value
                
                

                # Only fill if it's empty
                if not existing_teaching:
                    print(f"Fetching teachings for {coach_name} on {topic}...")
                    prompt = f"Question: Tell me what you know about {coach_name}'s teachings on {topic}. This includes any advice, tips, or strategies."
                    
                    
                    #  Save both full and summary embeddings in the vector database
                    vector_id = create_ascii_vector_id(coach_name, topic)

                    # Check if the vector ID already exists in the index
                    existing_ids = list(index_full.fetch([vector_id])['vectors'].keys())

                    if existing_ids:                   
                        print(f"Embedding for {vector_id} already exists in the {index_full} index. Skipping upsert.")
                    else:
                        
                        # Get teachings from Vertex AI
                        teachings = generate_response(prompt)
                        #print(teachings)


                        # 4a. Generate embeddings for the full teachings
                        full_embedding = embeddings.encode(teachings)

                        # 4b. Generate a summary of the teachings
                        summary = summarizer(teachings, max_length=50, min_length=25, do_sample=False)[0]['summary_text']
                            
                        # 4c. Generate embeddings for the summary
                        summary_embedding = embeddings.encode(summary)



                        # Upsert full document embedding to the full index
                        upsert_with_notification(index_full,vector_id, full_embedding, "full")
                            

                         # Upsert summary embedding to the summary index
                        upsert_with_notification(index_summary,vector_id, summary_embedding, "summary")
                        
                        # Fill the teachings in the cell
                        sheet.cell(row=row, column=col).value = teachings
  

    # Save the updated Excel workbook
    wb.save(file_path)
    print("Excel sheet updated successfully.")
    return file_path
# 6a. Helper function to create a clean ASCII vector ID
def create_ascii_vector_id(coach_name, topic):
    # Create an ID by combining coach_name and topic, then remove non-ASCII characters
    raw_id = f"{coach_name}_{topic}".replace(" ", "_")
    clean_id = re.sub(r'[^\x00-\x7F]+', '', raw_id)  # Remove non-ASCII characters
    return clean_id

# 6b. Helper function to upsert an embedding into an index and notify when successful
def upsert_with_notification(index, vector_id, embedding, index_type):
    """
    Upserts an embedding into the specified Pinecone index and notifies when successful.

    Parameters:
    - index (str): The  Pinecone index (either full or summary).
    - vector_id (str): Unique ID for the vector being upserted.
    - embedding (list): The embedding vector to be upserted.
    - index_type (str): Type of the index (either 'full' or 'summary') for notifications.
    """
   

    # Upsert the embedding into the specified index
    response = index.upsert([(vector_id, embedding)])

    # Check if the upsert was successful (based on Pinecone's response or wait time)
    if response:
        print(f"Embedding for {vector_id} successfully upserted into the {index_type} index!")
    else:
        print(f"Failed to upsert embedding for {vector_id} into the {index_type} index.")
    

# 7. Create the chatbot function
def create_chatbot():
    index_name = "life-coaches-conversation"
    create_index_if_not_exists(index_name)
    index = pc.Index(index_name)

    vectorstore = PineconeVectorStore(index=index)
    
    chatbot = RetrievalQA.from_chain_type(
        llm=VertexAI(),
        chain_type="stuff",
        retriever=vectorstore.as_retriever(search_kwargs={"k": 3})
    )
    return chatbot



# 8. SET UP VERTEX AI MODEL (FOR GENERATION)
# Function to generate text using Vertex AI model
@retry(wait=wait_exponential(multiplier=1, min=4, max=10), stop=stop_after_attempt(5), retry=retry_if_exception_type(ResourceExhausted))
def generate_response(prompt):
    
    # Fallback response for error handling
    empty_response = RunnableLambda(
        lambda x: AIMessage(content="Error processing document")
    )

    # Initialize the Vertex AI model client
    vertex_ai_client = VertexAI(
    temperature=1, model_name="gemini-1.0-pro", max_tokens=1024
    ).with_fallbacks([empty_response])

    # Define the input for the LLMChain
    var = {
       "text": prompt
    }

    time.sleep(30)
    # Create and run the LLMChain
    chain = LLMChain(
        llm=vertex_ai_client,
        prompt=PromptTemplate(
            input_variables=["text"],
            template=prompt
        )
    )
    response = chain.run(var)  
    #return response.predictions[0]
    
    return response

# 9. Modified RAG-based response generation function
def generate_rag_response(question):
    # Encode the query with SentenceTransformer
    query_embedding = embeddings.encode(question).tolist()  # Convert to list for compatibility

    # Retrieve context from Pinecone
    index = pc.Index(index_name_full)  # Use the Index object for similarity search
    retrievals = index.query(vector=query_embedding, top_k=3, include_metadata=True)  # Perform query
    
    # Check for metadata in matches and handle missing metadata gracefully
    context_parts = []
    for match in retrievals["matches"]:
        # Attempt to access 'metadata' and get 'text', or add a fallback if not present
        text_content = match.get("metadata", {}).get("text", "[No relevant text found in metadata]")
        context_parts.append(text_content)
    
    # Combine retrieved texts for context
    context = " ".join(context_parts)
    
    # Generate response with the Vertex AI model using the context
    prompt = f"Question: {question}\nContext: {context}\nAnswer:"
    response = generate_response(prompt)
    
    return response

# Check if user profile exists in Pinecone
def get_user_profile(phone_number):
    index_name_profile = "user-profile"
    index = create_index_if_not_exists(index_name_profile)
    profile_id = f"profile_{phone_number}"
    
    try:
        # Fetch directly by ID
        result = index.fetch(ids=[profile_id])
        print("Fetch result:", result)

        # Check if profile exists in the result
        if result and profile_id in result['vectors']:
            return result['vectors'][profile_id]["metadata"]
    except Exception as e:
        print("Error during fetch:", e)

    print("Profile not found or fetch error.")
    return None

# Save user profile metadata to Pinecone
def save_user_profile(phone_number, metadata):
    index_name_profile = "user-profile"
    index = create_index_if_not_exists(index_name_profile)
    
    profile_id = f"profile_{phone_number}"
    # Create embeddings for the user profile metadata
    metadata_text = f"{metadata['name']} {metadata['goals']} {metadata['focus_area']}"
    profile_embedding = embeddings.encode(metadata_text).tolist()

    # Upsert the profile with embeddings into Pinecone
    index.upsert([(profile_id, profile_embedding, metadata)])
    print(f"User profile for {phone_number} saved.")




# 10. FLASK ROUTE FOR LOADING KNOWLEDGE BASE
app = Flask(__name__)

# Temporary in-memory storage for chat history
chat_history = {}
user_states = {}


# List of initial questions for new users
initial_questions = [
    "What's your name?",
    "What are your main goals for coaching?",
    "Any specific areas you'd like help with (e.g., career, personal growth)?"
]

@app.route('/')
def index():
    return render_template('index.html')
# Define a route to load the knowledge base
@app.route('/load-knowledge-base', methods=['POST'])
def load_knowledge_base_route():
    try: 
        documents = load_knowledge_base()
        print("Knowledge base loaded:")
        return jsonify({"documents": documents})
    except HTTPError as e:
        return jsonify({"error": str(e)}), 404
# Route to start chat and check for first-time users
@app.route("/start", methods=["POST"])
def start_chat():
    data = request.get_json()
    phone_number = data.get("phone_number")

    if not phone_number:
        return jsonify({"error": "Phone number is required"}), 400

    # Check if user profile exists
    user_profile = get_user_profile(phone_number)

    # If profile exists, greet user and return to chat
    if user_profile:
        return jsonify({
            "message":  f"Welcome back, {user_profile['name']}!",
            "new_user": False
            })
    else: 
        print("New user detected, sending initial questions.")
        # Return the first question
        return jsonify({
            "message": "Hello! Thank you for reaching out. Before we begin, I will need your approval to collect and store your responses to the following questions. Your information will be kept confidential and used only for the purpose of providing coaching services. Are you ready to proceed? (Yes/No)",
            "new_user": True,
        })

# Endpoint to collect user responses to initial questions
@app.route("/collect_initial_data", methods=["POST"])
def collect_initial_data():
    try:
        # Parse JSON data from request
        data = request.get_json()

        # Extract data fields
        phone_number = data.get("phone_number", "")
        approval = data.get("approval", "")
        name = data.get("name", "")
        goals = data.get("goals", "")
        focus_area = data.get("focus_area", "")

        # Construct metadata
        user_metadata = {
            "phone_number": phone_number,
            "approval": approval,
            "name": name,
            "goals": goals,
            "focus_area": focus_area,
            "timestamp": datetime.now().isoformat()
        }
        print("Received user metadata:", user_metadata)

        # Save user profile - ensure `save_user_profile` is set up to accept metadata without a vector if needed
        save_user_profile(phone_number, user_metadata)

        # Send success response
        return jsonify({"message": f"Thanks, {name}! We’re ready to start."}), 200

    except Exception as e:
        print("Error in collect_initial_data:", e)
        return jsonify({"error": "An error occurred while processing data."}), 500

# Function to retrieve coach based on user needs
def retrieve_coach_based_on_user_needs(focus_area, goals, index_summary):
    """
    Finds the top 3 coaches that best match the user's focus area and goals by querying summary_index.
    
    Parameters:
    - focus_area (str): The main area of interest or focus for the user.
    - goals (str): Specific goals the user has in mind.
    - index_summary (pinecone.Index): The Pinecone index for coach summaries.

    Returns:
    - List[dict]: A list of the top 3 matching coaches with their metadata.
    """
    
    # Combine focus area and goals to create the query text
    query_text = f"{focus_area} {goals}"
    
    # Generate the embedding for the query
    query_embedding = embeddings.encode(query_text).tolist()  # Convert to list for Pinecone compatibility
    
    # Perform the query on the summary_index for top 3 results
    response = index_summary.query(vector=query_embedding, top_k=3, include_metadata=True)
    
    # Collect and format the top 3 coaches
    top_coaches = []
    for match in response["matches"]:
        # Extract metadata from the match and structure it
        coach_data = match["metadata"]
        top_coaches.append({
            "name": coach_data.get("name"),
            "specialization": coach_data.get("specialization"),
            "experience": coach_data.get("experience_years", "N/A"),
            "bio": coach_data.get("bio", "No bio available"),
            "score": match.get("score")  # Optionally include the similarity score
        })
        
    # If no matches found, return a default message
    if not top_coaches:
        return [{"message": "No suitable coaches found. Please adjust focus area or goals."}]
    
    return top_coaches
    

# Endpoint for continuing the chat with the bot after initialization
@app.route("/chat", methods=["POST"])
def chat():
    data = request.get_json()
    phone_number = data.get('phone_number')
    question = data.get('question')

    if not phone_number or not question:
        return jsonify({"error": "Phone number and question are required"}), 400

    # Process the question and generate a response
    response = generate_rag_response(question)

    # Save the chat history
    return jsonify({"message": response})

# Optional: Route to get chat history for a user
@app.route("/history/<phone_number>", methods=["GET"])
def get_chat_history(phone_number):
    if phone_number in chat_history:
        return jsonify(chat_history[phone_number])
    else:
        return jsonify({"error": "No conversation found for this number"}), 404

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)


