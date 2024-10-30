import openpyxl
import re
import time
from managers.chatbot import Chatbot

class KnowledgeBaseManager:
    def __init__(self, embeddings, summarizer, pinecone_manager):
        self.embeddings = embeddings
        self.summarizer = summarizer
        self.pinecone_manager = pinecone_manager
        self.index_name_full = "life-coaches-full"
        self.index_name_summary = "life-coaches-summary"
        self.index_full = self.pinecone_manager.create_index(self.index_name_full, self.embeddings.get_sentence_embedding_dimension())
        self.index_summary = self.pinecone_manager.create_index(self.index_name_summary, self.embeddings.get_sentence_embedding_dimension())

    def load_knowledge_base(self, excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            coach_name = sheet[f"A{row}"].value
            for col in range(4, sheet.max_column + 1):
                topic = sheet.cell(row=1, column=col).value
                existing_teaching = sheet.cell(row=row, column=col).value
                if topic and not existing_teaching:
                    prompt = f"Question: Tell me about {coach_name}'s teachings on {topic}."
                    vector_id = self.create_vector_id(coach_name, topic)
                    if not self.index_full.fetch([vector_id])['vectors']:
                        teachings = Chatbot().generate_response(prompt)
                        full_embedding = self.embeddings.encode(teachings)
                        summary = self.summarizer(teachings, max_length=50, min_length=25, do_sample=False)[0]['summary_text']
                        summary_embedding = self.embeddings.encode(summary)
                        self.upsert_embedding(self.index_full, vector_id, full_embedding, "full")
                        self.upsert_embedding(self.index_summary, vector_id, summary_embedding, "summary")
                        sheet.cell(row=row, column=col).value = teachings
        wb.save(excel_file_path)
        print("Knowledge base loaded successfully.")

    @staticmethod
    def create_vector_id(coach_name, topic):
        raw_id = f"{coach_name}_{topic}".replace(" ", "_")
        return re.sub(r'[^\x00-\x7F]+', '', raw_id)

    def upsert_embedding(self, index, vector_id, embedding, index_type):
        response = index.upsert([(vector_id, embedding)])
        print(f"Upsert for {vector_id} into {index_type} index: {'successful' if response else 'failed'}")
